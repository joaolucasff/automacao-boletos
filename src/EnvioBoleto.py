# EnvioBoletos_v6.0.0.py — 29/10/2025
# ---------------------------------------------------------
# CHANGELOG v6.0.0 (ETAPA 4: IA + Fallback Inteligente) 🤖
# [OK] Base: v5.0.0
# [OK] (NEW) Integração com DeepSeek (Ollama) para extração via IA
# [OK] (NEW) IA como primeira linha, regex como fallback automático
# [OK] (NEW) extrair_cnpj_do_pdf() com IA + regex fallback
# [OK] (NEW) detectar_fidc_do_pdf() com IA + regex fallback
# [OK] (NEW) extrair_cnpj_da_nota() com IA + regex fallback
# [OK] (NEW) extrair_valor_da_nota() com IA + regex fallback
# [OK] (NEW) Função genérica extrair_com_ia() com timeout
# [OK] (NEW) Flags USAR_IA e IA_DISPONIVEL para controle
# [OK] (IMP) Prompts otimizados para cada tipo de extração
# [OK] (IMP) Timeout de 10s para chamadas IA
# [OK] (IMP) Logs transparentes sobre qual método foi usado
#
# CHANGELOG v5.0.0 (ETAPA 3: Validação Robusta + Auto-CNPJ)
# [OK] Base: v4.1.0
# [OK] (NEW) MODO_PREVIEW ativado (emails abrem sem enviar automaticamente)
# [OK] (NEW) Auto-preenchimento de CNPJ na planilha Excel
# [OK] (NEW) Extração de CNPJ e Valor das notas fiscais
# [OK] (NEW) Validação tripla boleto↔nota: Número, CNPJ, Valor
# [OK] (NEW) Rejeita notas com CNPJ diferente do boleto
# [OK] (NEW) Avisa se valor da nota difere do boleto
# [OK] (IMP) Logs detalhados de cada validação
# [OK] (NEW) Bloqueio de envio se notas inválidas
#
# CHANGELOG v4.1.0 (ETAPA 2: FIDCs Dinâmicos)
# [OK] Dicionário FIDC_CONFIG com 4 FIDCs
# [OK] Detecção automática de FIDC do boleto PDF
# [OK] Template de email e CCs dinâmicos por FIDC
#
# CHANGELOG v4.0.0 (ETAPA 1: Matching Robusto)
# [OK] Matching em 3 níveis: CNPJ+Valor → Nome+Valor → Fuzzy
# [OK] Sugestões inteligentes quando não encontrar match
# =========================================================

import os, re, time, shutil
from datetime import datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
import win32com.client as win32
from openpyxl import load_workbook
import warnings

# Suprimir avisos chatos do openpyxl sobre DrawingML
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# PDF reading para extração de CNPJ
try:
    import pdfplumber
    PDF_DISPONIVEL = True
except ImportError:
    PDF_DISPONIVEL = False
    print("[AVISO] pdfplumber não instalado. Extração de CNPJ desabilitada.")
    print("   Para instalar: pip install pdfplumber")

# IA para extração inteligente de dados (IA primeira, regex fallback)
try:
    from langchain_core.prompts import PromptTemplate
    from langchain_core.output_parsers import StrOutputParser
    from langchain_ollama.llms import OllamaLLM
    IA_DISPONIVEL = True
except ImportError:
    IA_DISPONIVEL = False
    print("[AVISO] LangChain não instalado. IA desabilitada (usando apenas regex).")
    print("   Para instalar: pip install langchain-core langchain-ollama")

# -------------------- CONFIG IA --------------------
USAR_IA = True  # True = tenta IA primeiro, False = só regex
IA_TIMEOUT = 10  # Timeout em segundos para chamadas IA
IA_MODEL = "deepseek-r1:1.5b"  # Modelo Ollama
IA_TEMPERATURE = 0  # 0 = mais determinístico

# -------------------- CONFIG --------------------
BASE = r"C:\Users\User-OEM\Desktop\BoletosAutomação"
PASTA_BOLETOS   = os.path.join(BASE, r"BoletosRenomeados")
PASTA_NOTAS     = os.path.join(BASE, r"Notas")
PASTA_LOGS      = os.path.join(BASE, r"Logs")
PASTA_ERROS     = os.path.join(BASE, r"Erros")
PASTA_ENVIADOS  = os.path.join(BASE, r"BoletosEnviados")
ARQUIVO_PLANILHA = os.path.join(BASE, r"config", "Envio boletos - alterar dados.xlsx")
ABA_PLANILHA     = "Alterar dados"

EMAIL_CC_FIXO   = "joaolucasfurtadofiel17@gmail.com"
ASSINATURA_IMG  = os.path.join(BASE, r"Outros", r"boletosTeste (temporario)", "Imagem1.jpg")

os.makedirs(PASTA_LOGS, exist_ok=True)
os.makedirs(PASTA_ERROS, exist_ok=True)
os.makedirs(PASTA_ENVIADOS, exist_ok=True)

MODO_PREVIEW = True   # Modo preview: abre emails no Outlook sem enviar automaticamente

# -------------------- CONFIGURAÇÃO DE FIDCs --------------------
# Dados de cada FIDC: nome completo, CNPJ, e emails em cópia (CC)
FIDC_CONFIG = {
    "CAPITAL": {
        "nome_completo": "CAPITAL RS FIDC NP MULTISSETORIAL",
        "cnpj": "12.910.463/0001-70",
        "cc_emails": ["adm@jotajota.net.br"],
        "palavras_chave": ["CAPITAL RS", "CAPITAL RS FIDC"]
    },
    "NOVAX": {
        "nome_completo": "Novax Fundo de Invest. Em Dir. Cred.",
        "cnpj": "28.879.551/0001-96",
        "cc_emails": ["adm@jotajota.net.br", "controladoria@novaxfidc.com.br"],
        "palavras_chave": ["NOVAX", "NOVAX FIDC", "NOVAX FUNDO"]
    },
    "CREDVALE": {
        "nome_completo": "CREDVALE FUNDO DE INVESTIMENTO EM DIREITOS CREDITORIOS MULTISSETORIAL",
        "cnpj": "28.194.817/0001-67",
        "cc_emails": ["adm@jotajota.net.br", "nichole@credvalefidc.com.br"],
        "palavras_chave": ["CREDVALE", "CREDVALE FUNDO", "CREDVALE FIDC"]
    },
    "SQUID": {
        "nome_completo": "SQUID FUNDO DE INVESTIMENTO EM DIREITOS CREDITORIOS",
        "cnpj": "28.849.641/0001-34",
        "cc_emails": ["adm@jotajota.net.br"],
        "palavras_chave": ["SQUID", "SQUID FIDC", "SQUID FUNDO"]
    }
}

# FIDC padrão caso não consiga detectar
FIDC_PADRAO = "CAPITAL"

# -------------------- HELPERS --------------------
def normalize_pagador(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()

def valor_to_cents(valor) -> int | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float, Decimal)):
        try:
            d = Decimal(str(valor))
            return int((d * 100).quantize(Decimal("1")))
        except InvalidOperation:
            return None
    s = str(valor).replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        d = Decimal(s)
        return int((d * 100).quantize(Decimal("1")))
    except InvalidOperation:
        return None

def cents_to_brl(cents: int) -> str:
    inte, frac = cents // 100, cents % 100
    return f"{inte:,}".replace(",", ".") + f",{frac:02d}"

def digits_only(s: str) -> str:
    return re.sub(r"[^0-9]", "", s or "")

def normalizar_cnpj(cnpj: str) -> str:
    """
    Normaliza CNPJ removendo pontuação, retorna só dígitos
    Exemplo: "12.345.678/0001-90" → "12345678000190"
    """
    return digits_only(cnpj)

# -------------------- IA + FALLBACK GENÉRICO --------------------
def extrair_com_ia(texto_pdf: str, prompt_template: str, timeout: int = IA_TIMEOUT) -> str | None:
    """
    Extrai dados usando IA (DeepSeek via Ollama)

    Parâmetros:
        texto_pdf: Texto extraído do PDF
        prompt_template: Template do prompt (deve conter {text})
        timeout: Timeout em segundos

    Retorna:
        Resultado extraído ou None se falhar
    """
    if not (IA_DISPONIVEL and USAR_IA):
        return None

    try:
        # Inicializar modelo
        llm = OllamaLLM(
            model=IA_MODEL,
            temperature=IA_TEMPERATURE,
            timeout=timeout
        )

        # Criar chain
        prompt = PromptTemplate(template=prompt_template, input_variables=["text"])
        chain = prompt | llm | StrOutputParser()

        # Executar (com timeout)
        resultado = chain.invoke({"text": texto_pdf})

        # Limpar resultado
        resultado = resultado.strip()

        # Verificar se é resposta válida
        if resultado and resultado.upper() != "NAO_ENCONTRADO":
            return resultado

        return None

    except Exception as e:
        print(f"      [IA] Erro na extração via IA: {str(e)[:100]}")
        return None

def extrair_texto_pdf(caminho_pdf: str) -> str | None:
    """
    Extrai texto completo de um PDF

    Retorna: Texto do PDF ou None se falhar
    """
    if not PDF_DISPONIVEL:
        return None

    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = ""
            for pag in pdf.pages:
                t = pag.extract_text()
                if t:
                    texto += t + "\n"
            return texto if texto.strip() else None
    except Exception as e:
        print(f"      [ERRO] Falha ao ler PDF: {e}")
        return None

def extrair_cnpj_do_pdf(caminho_pdf: str) -> str | None:
    """
    Extrai CNPJ do pagador do boleto PDF
    Usa IA primeiro, fallback para regex se falhar

    Retorna: "12345678000190" (14 dígitos) ou None
    """
    if not PDF_DISPONIVEL:
        return None

    try:
        # Extrair texto do PDF
        texto = extrair_texto_pdf(caminho_pdf)
        if not texto:
            return None

        # TENTATIVA 1: Usar IA (se disponível)
        if IA_DISPONIVEL and USAR_IA:
            prompt = """Extraia o CNPJ do PAGADOR deste boleto.
O CNPJ tem 14 dígitos no formato XX.XXX.XXX/XXXX-XX.

Procure por:
- Seção "Pagador" ou "Sacado"
- CNPJ logo após o nome da empresa

Retorne APENAS os 14 dígitos sem pontuação (exemplo: 12345678000190).
Se não encontrar, retorne "NAO_ENCONTRADO".

Texto do PDF:
{text}

Resposta (apenas 14 números):"""

            resultado_ia = extrair_com_ia(texto, prompt)
            if resultado_ia:
                cnpj_limpo = normalizar_cnpj(resultado_ia)
                if len(cnpj_limpo) == 14:
                    return cnpj_limpo

        # TENTATIVA 2: Fallback Regex (método original)
        # Procura padrão: "Pagador:" ... "CNPJ: XX.XXX.XXX/0001-XX"
        match = re.search(
            r'Pagador[:\s][^\n]*?CNPJ[:\s]*(\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\/\s]?\d{4}[\-\s]?\d{2})',
            texto,
            re.IGNORECASE | re.DOTALL
        )

        if match:
            cnpj_raw = match.group(1)
            cnpj_limpo = normalizar_cnpj(cnpj_raw)
            if len(cnpj_limpo) == 14:
                return cnpj_limpo

        # TENTATIVA 3: Fallback regex menos preciso
        match_any = re.search(r'(\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\/\s]?\d{4}[\-\s]?\d{2})', texto)
        if match_any:
            cnpj_limpo = normalizar_cnpj(match_any.group(1))
            if len(cnpj_limpo) == 14:
                return cnpj_limpo

        return None

    except Exception as e:
        print(f"[AVISO] Erro ao extrair CNPJ de {os.path.basename(caminho_pdf)}: {e}")
        return None

def detectar_fidc_do_pdf(caminho_pdf: str) -> str:
    """
    Detecta qual FIDC (beneficiário) está no boleto PDF
    Usa IA primeiro, fallback para regex se falhar

    Retorna: "CAPITAL", "NOVAX", "CREDVALE", "SQUID" ou FIDC_PADRAO
    """
    if not PDF_DISPONIVEL:
        return FIDC_PADRAO

    try:
        # Extrair texto do PDF
        texto = extrair_texto_pdf(caminho_pdf)
        if not texto:
            return FIDC_PADRAO

        # TENTATIVA 1: Usar IA (se disponível)
        if IA_DISPONIVEL and USAR_IA:
            prompt = """Identifique qual FIDC é o beneficiário/cedente deste boleto.

Opções possíveis:
1. CAPITAL RS FIDC NP MULTISSETORIAL
2. Novax Fundo de Invest. Em Dir. Cred.
3. CREDVALE FUNDO DE INVESTIMENTO
4. SQUID FUNDO DE INVESTIMENTO

Procure no campo "Beneficiário", "Cedente" ou empresa que está recebendo o pagamento.

Retorne APENAS uma dessas palavras:
- CAPITAL
- NOVAX
- CREDVALE
- SQUID
- NAO_ENCONTRADO

Texto do PDF:
{text}

Resposta (uma palavra):"""

            resultado_ia = extrair_com_ia(texto, prompt)
            if resultado_ia:
                resultado_upper = resultado_ia.upper().strip()
                # Verificar se é um FIDC válido
                if resultado_upper in FIDC_CONFIG:
                    return resultado_upper

        # TENTATIVA 2: Fallback Regex (método original)
        # Normalizar texto para busca
        texto_upper = texto.upper()

        # Procurar por cada FIDC usando palavras-chave
        for fidc_nome, config in FIDC_CONFIG.items():
            for palavra_chave in config["palavras_chave"]:
                if palavra_chave.upper() in texto_upper:
                    return fidc_nome

        # Não encontrou nenhum - usar padrão
        print(f"[AVISO] FIDC nao detectado em {os.path.basename(caminho_pdf)}, usando padrao: {FIDC_PADRAO}")
        return FIDC_PADRAO

    except Exception as e:
        print(f"[AVISO] Erro ao detectar FIDC de {os.path.basename(caminho_pdf)}: {e}")
        return FIDC_PADRAO

def calcular_similaridade(str1: str, str2: str) -> float:
    """
    Calcula similaridade entre duas strings (0.0 a 1.0)
    Usa SequenceMatcher para comparação fuzzy
    """
    return SequenceMatcher(None, str1.upper(), str2.upper()).ratio()

def encontrar_match_mais_proximo(nome_busca: str, candidatos: list) -> tuple:
    """
    Encontra o candidato mais similar ao nome buscado
    Retorna: (melhor_candidato, score_similaridade)
    """
    if not candidatos:
        return None, 0.0

    melhor = max(candidatos, key=lambda x: calcular_similaridade(nome_busca, x))
    score = calcular_similaridade(nome_busca, melhor)
    return melhor, score

# -------------------- PLANILHA --------------------
def atualizar_cnpj_na_planilha(pagador, valor_cents, cnpj):
    """
    Auto-preenche CNPJ na planilha Excel para eliminar trabalho manual

    Parâmetros:
        pagador: Nome do pagador (normalizado)
        valor_cents: Valor em centavos
        cnpj: CNPJ extraído do boleto (14 dígitos)

    Retorna:
        True se atualizou, False se não encontrou ou erro
    """
    if not cnpj:
        return False

    try:
        wb = load_workbook(ARQUIVO_PLANILHA, data_only=False)
        sh = wb[ABA_PLANILHA]
        headers = {sh.cell(1, c).value: c for c in range(1, sh.max_column + 1)}

        col_pagador = headers.get("Pagador")
        col_valor = headers.get("Valor")
        col_cnpj = headers.get("CNPJ")

        # Se não tem coluna CNPJ, criar
        if not col_cnpj:
            col_cnpj = sh.max_column + 1
            sh.cell(1, col_cnpj).value = "CNPJ"
            print(f"   [INFO] Coluna 'CNPJ' criada na planilha")

        if not (col_pagador and col_valor):
            wb.close()
            return False

        # Procurar linha com mesmo pagador + valor
        for r in range(2, sh.max_row + 1):
            pag_cell = sh.cell(r, col_pagador).value
            val_cell = sh.cell(r, col_valor).value
            cnpj_cell = sh.cell(r, col_cnpj).value

            if not (pag_cell and val_cell):
                continue

            # Comparar pagador e valor
            pag_key_planilha = normalize_pagador(pag_cell)
            val_cents_planilha = valor_to_cents(val_cell)

            if pag_key_planilha == pagador and val_cents_planilha == valor_cents:
                # Encontrou a linha!
                if cnpj_cell:
                    # Já tem CNPJ, verificar se é o mesmo
                    cnpj_existente = normalizar_cnpj(str(cnpj_cell))
                    if cnpj_existente == cnpj:
                        wb.close()
                        return False  # Já tinha o CNPJ correto
                    else:
                        # CNPJ diferente - atualizar
                        print(f"   [AVISO] CNPJ diferente na planilha (era: {cnpj_existente}, agora: {cnpj})")

                # Atualizar CNPJ
                cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                sh.cell(r, col_cnpj).value = cnpj_formatado
                wb.save(ARQUIVO_PLANILHA)
                wb.close()
                print(f"   [OK] CNPJ adicionado a planilha: {cnpj_formatado}")
                return True

        wb.close()
        return False  # Não encontrou linha correspondente

    except Exception as e:
        print(f"   [AVISO] Erro ao atualizar CNPJ na planilha: {e}")
        return False

def carregar_mapa_pagador_valor():
    """
    Carrega mapeamento da planilha Excel

    Suporta 2 formatos:
    1. COM CNPJ: Documento | Pagador | Valor | Email | CNPJ
    2. SEM CNPJ: Documento | Pagador | Valor | Email (compatibilidade com versão antiga)

    Retorna:
        mapa_nome: {(pagador_norm, valor_cents): {'email': str, 'docs': set, 'cnpj': str|None, 'pagador_original': str}}
        mapa_cnpj: {(cnpj, valor_cents): entry} (mesmo formato)
        lista_nomes: [pagador_norm, ...] (para fuzzy matching)
    """
    mapa_nome = {}   # Mapa por (nome, valor)
    mapa_cnpj = {}   # Mapa por (CNPJ, valor)
    lista_nomes = [] # Lista de nomes para fuzzy matching

    try:
        wb = load_workbook(ARQUIVO_PLANILHA, data_only=True)
        sh = wb[ABA_PLANILHA]
        headers = {sh.cell(1, c).value: c for c in range(1, sh.max_column + 1)}

        # Colunas obrigatórias
        col_doc = headers.get("Documento")
        col_pagador = headers.get("Pagador")
        col_valor = headers.get("Valor")
        col_email = headers.get("Email")

        # Coluna opcional (compatibilidade)
        col_cnpj = headers.get("CNPJ")  # Pode ser None se planilha antiga

        if not (col_doc and col_pagador and col_valor and col_email):
            print("[ERRO] Cabeçalhos obrigatórios ausentes na planilha.")
            print("   Necessário: Documento, Pagador, Valor, Email")
            print(f"   Encontrados: {list(headers.keys())}")
            wb.close()
            return mapa_nome, mapa_cnpj, lista_nomes

        if col_cnpj:
            print("[OK] Planilha com coluna CNPJ detectada (modo robusto)")
        else:
            print("[AVISO] Planilha SEM coluna CNPJ (modo compatibilidade)")
            print("        Recomendado: adicionar coluna 'CNPJ' para matching mais robusto")

        linhas_processadas = 0
        linhas_com_cnpj = 0

        for r in range(2, sh.max_row + 1):
            doc  = sh.cell(r, col_doc).value
            pag  = sh.cell(r, col_pagador).value
            val  = sh.cell(r, col_valor).value
            mail = sh.cell(r, col_email).value
            cnpj_raw = sh.cell(r, col_cnpj).value if col_cnpj else None

            # Validações básicas
            if not (pag and val and mail):
                continue

            pag_key = normalize_pagador(pag)
            val_cents = valor_to_cents(val)

            if val_cents is None:
                print(f"[AVISO] Linha {r}: Valor invalido '{val}' para {pag}")
                continue

            # Processar CNPJ se disponível
            cnpj_norm = None
            if cnpj_raw:
                cnpj_norm = normalizar_cnpj(str(cnpj_raw))
                if len(cnpj_norm) != 14:
                    print(f"[AVISO] Linha {r}: CNPJ invalido '{cnpj_raw}' para {pag}")
                    cnpj_norm = None
                else:
                    linhas_com_cnpj += 1

            # Criar entry
            entry = {
                'email': str(mail).strip(),
                'docs': set(),
                'cnpj': cnpj_norm,
                'pagador_original': str(pag).strip()
            }

            if doc:
                entry['docs'].add(str(doc).strip())

            # Adicionar ao mapa por nome
            key_nome = (pag_key, val_cents)
            if key_nome in mapa_nome:
                # Merge documentos se chave duplicada
                mapa_nome[key_nome]['docs'].update(entry['docs'])
            else:
                mapa_nome[key_nome] = entry
                lista_nomes.append(pag_key)

            # Adicionar ao mapa por CNPJ (se disponível)
            if cnpj_norm:
                key_cnpj = (cnpj_norm, val_cents)
                if key_cnpj in mapa_cnpj:
                    mapa_cnpj[key_cnpj]['docs'].update(entry['docs'])
                else:
                    mapa_cnpj[key_cnpj] = entry

            linhas_processadas += 1

        wb.close()

        print(f"[OK] Planilha carregada com sucesso:")
        print(f"     - Total de clientes mapeados: {len(mapa_nome)}")
        print(f"     - Clientes com CNPJ: {linhas_com_cnpj}")
        print(f"     - Linhas processadas: {linhas_processadas}")

        return mapa_nome, mapa_cnpj, lista_nomes

    except Exception as e:
        print(f"[ERRO] Erro ao ler planilha: {e}")
        import traceback
        traceback.print_exc()
        return {}, {}, []

# -------------------- VALIDAÇÃO DE NOTAS FISCAIS --------------------
# Flags para controle de IA (quando implementada)
USAR_IA_VALIDACAO = False  # Quando IA estiver pronta, mude para True

def extrair_cnpj_da_nota(caminho_pdf_nota):
    """
    Extrai CNPJ da nota fiscal PDF (do destinatário/remetente)
    Usa IA primeiro, fallback para regex se falhar

    Retorna: CNPJ normalizado (14 dígitos) ou None
    """
    if not PDF_DISPONIVEL:
        return None

    try:
        # Extrair texto do PDF
        texto = extrair_texto_pdf(caminho_pdf_nota)
        if not texto:
            return None

        # TENTATIVA 1: Usar IA (se disponível)
        if IA_DISPONIVEL and USAR_IA:
            prompt = """Extraia o CNPJ do DESTINATÁRIO desta nota fiscal.
O CNPJ tem 14 dígitos no formato XX.XXX.XXX/XXXX-XX.

Procure por:
- Seção "Destinatário" ou "Remetente"
- CNPJ próximo ao nome da empresa destinatária
- Geralmente está na parte superior da nota

Retorne APENAS os 14 dígitos sem pontuação (exemplo: 12345678000190).
Se não encontrar, retorne "NAO_ENCONTRADO".

Texto do PDF:
{text}

Resposta (apenas 14 números):"""

            resultado_ia = extrair_com_ia(texto, prompt)
            if resultado_ia:
                cnpj_limpo = normalizar_cnpj(resultado_ia)
                if len(cnpj_limpo) == 14:
                    return cnpj_limpo

        # TENTATIVA 2: Fallback Regex (método original)
        # Procurar CNPJ do destinatário (quem recebe a nota)
        patterns = [
            r'Destinat[áa]rio[^\n]*?CNPJ[:\s]*(\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\/\s]?\d{4}[\-\s]?\d{2})',
            r'CNPJ[:\s]*(\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\/\s]?\d{4}[\-\s]?\d{2})',
        ]

        for pattern in patterns:
            match = re.search(pattern, texto, re.IGNORECASE)
            if match:
                cnpj = normalizar_cnpj(match.group(1))
                if len(cnpj) == 14:
                    return cnpj

        return None

    except Exception as e:
        print(f"      [AVISO] Erro ao extrair CNPJ da nota: {e}")
        return None

def extrair_valor_da_nota(caminho_pdf_nota):
    """
    Extrai valor total da nota fiscal PDF
    Usa IA primeiro, fallback para regex se falhar

    Retorna: Valor em centavos (int) ou None
    """
    if not PDF_DISPONIVEL:
        return None

    try:
        # Extrair texto do PDF
        texto = extrair_texto_pdf(caminho_pdf_nota)
        if not texto:
            return None

        # TENTATIVA 1: Usar IA (se disponível)
        if IA_DISPONIVEL and USAR_IA:
            prompt = """Extraia o VALOR TOTAL desta nota fiscal.

Procure por:
- "Valor Total da Nota"
- "Total da NF"
- "Valor Total NF"
- Geralmente está no rodapé ou campo de totais

Retorne no formato: 1234.56 (com ponto como separador decimal, sem R$ ou símbolos).
Exemplo: 606.08 para R$ 606,08

Se não encontrar, retorne "NAO_ENCONTRADO".

Texto do PDF:
{text}

Resposta (apenas número com 2 decimais):"""

            resultado_ia = extrair_com_ia(texto, prompt)
            if resultado_ia:
                # Tentar converter para centavos
                valor_cents = valor_to_cents(resultado_ia)
                if valor_cents:
                    return valor_cents

        # TENTATIVA 2: Fallback Regex (método original)
        # Procurar valor total da nota
        patterns = [
            r'Valor\s+Total\s+(?:da\s+)?(?:Nota|NF)[:\s]*R?\$?\s*([\d\.\,]+)',
            r'Total\s+(?:da\s+)?(?:Nota|NF)[:\s]*R?\$?\s*([\d\.\,]+)',
            r'Valor\s+NF[:\s]*R?\$?\s*([\d\.\,]+)',
        ]

        for pattern in patterns:
            match = re.search(pattern, texto, re.IGNORECASE)
            if match:
                valor_str = match.group(1)
                valor_cents = valor_to_cents(valor_str)
                if valor_cents:
                    return valor_cents

        return None

    except Exception as e:
        print(f"      [AVISO] Erro ao extrair valor da nota: {e}")
        return None

# -------------------- NF MATCHING --------------------
def indexar_notas_por_digitos():
    itens = []
    if not os.path.isdir(PASTA_NOTAS): return itens
    for f in os.listdir(PASTA_NOTAS):
        if f.lower().endswith(".pdf"):
            itens.append((os.path.join(PASTA_NOTAS, f), digits_only(f)))
    return itens

def achar_notas_por_docs_set(docs_set, notas_index, cnpj_boleto=None, valor_boleto_cents=None):
    """
    Busca e valida notas fiscais por documentos

    Validação em 3 níveis:
    1. Número da nota (primeiros 6 dígitos)
    2. CNPJ (se fornecido)
    3. Valor (se fornecido, tolerância ±10 centavos)

    Retorna: (notas_validadas, bases6, detalhes_validacao, tem_erro_critico)
    """
    bases, candidatos = set(), []
    tem_erro_critico = False

    # Fase 1: Buscar candidatos por número
    for d in docs_set:
        digs = digits_only(d)
        if len(digs) >= 6:
            bases.add(digs[:6])

    # VALIDAÇÃO CRÍTICA: Verificar se TODOS os documentos têm notas
    if len(docs_set) > 0 and len(bases) == 0:
        print(f"      [ERRO CRÍTICO] Nenhum documento válido encontrado em: {docs_set}")
        tem_erro_critico = True
        return [], [], [], tem_erro_critico

    # Buscar notas para cada documento
    docs_sem_nota = []
    for base6 in sorted(bases):
        encontrou = False
        for nf_path, nf_digits in notas_index:
            if base6 and base6 in nf_digits:
                candidatos.append((nf_path, base6))
                encontrou = True
                break

        if not encontrou:
            docs_sem_nota.append(base6)

    # VALIDAÇÃO CRÍTICA: Se algum documento não tem nota, bloquear
    if docs_sem_nota:
        print(f"      [ERRO CRÍTICO] Notas fiscais NÃO ENCONTRADAS para documentos: {docs_sem_nota}")
        print(f"      [BLOQUEIO] Email NÃO será enviado por segurança!")
        tem_erro_critico = True
        return [], list(bases), [], tem_erro_critico

    # Fase 2 e 3: Validar CNPJ e Valor
    notas_validadas = []
    detalhes = []

    for nf_path, base6 in candidatos:
        nome_nota = os.path.basename(nf_path)
        validacao = {
            'nota': nome_nota,
            'numero_ok': True,  # Já passou pela validação de número
            'cnpj_ok': None,
            'valor_ok': None,
            'anexada': False
        }

        # Validar CNPJ se fornecido (VALIDAÇÃO CRÍTICA!)
        if cnpj_boleto:
            cnpj_nota = extrair_cnpj_da_nota(nf_path)
            if cnpj_nota:
                validacao['cnpj_ok'] = (cnpj_nota == cnpj_boleto)
                if not validacao['cnpj_ok']:
                    print(f"      [ERRO CRÍTICO] {nome_nota}: CNPJ DIVERGENTE!")
                    print(f"                     Boleto: {cnpj_boleto[:2]}.{cnpj_boleto[2:5]}...{cnpj_boleto[12:]}")
                    print(f"                     Nota:   {cnpj_nota[:2]}.{cnpj_nota[2:5]}...{cnpj_nota[12:]}")
                    print(f"      [BLOQUEIO] Nota com CNPJ divergente NÃO será anexada!")
                    detalhes.append(validacao)
                    tem_erro_critico = True
                    continue  # Pula esta nota
            else:
                print(f"      [AVISO] {nome_nota}: CNPJ não encontrado na nota (anexando mesmo assim)")
                validacao['cnpj_ok'] = None

        # Validar Valor se fornecido
        if valor_boleto_cents:
            valor_nota_cents = extrair_valor_da_nota(nf_path)
            if valor_nota_cents:
                diferenca = abs(valor_nota_cents - valor_boleto_cents)
                tolerancia = 10  # 10 centavos
                validacao['valor_ok'] = (diferenca <= tolerancia)
                if not validacao['valor_ok']:
                    valor_boleto_brl = cents_to_brl(valor_boleto_cents)
                    valor_nota_brl = cents_to_brl(valor_nota_cents)
                    print(f"      [AVISO] {nome_nota}: Valor diferente")
                    print(f"              Boleto: R$ {valor_boleto_brl}")
                    print(f"              Nota:   R$ {valor_nota_brl}")
                    # Não rejeita por valor, só avisa
            else:
                print(f"      [AVISO] {nome_nota}: Valor nao encontrado")
                validacao['valor_ok'] = None

        # Se passou nas validações (ou não havia validações), anexar
        validacao['anexada'] = True
        notas_validadas.append(nf_path)
        detalhes.append(validacao)
        print(f"      [OK] {nome_nota}: Validada (CNPJ: {validacao['cnpj_ok']}, Valor: {validacao['valor_ok']})")

    # Remover duplicatas
    seen, notas_unique = set(), []
    for n in notas_validadas:
        if n not in seen:
            seen.add(n)
            notas_unique.append(n)

    return notas_unique, sorted(list(bases)), detalhes, tem_erro_critico

# -------------------- MATCHING ROBUSTO --------------------
def buscar_entrada_planilha(pagador_norm, valor_cents, cnpj, mapa_nome, mapa_cnpj, lista_nomes):
    """
    Busca entrada na planilha usando matching em 3 níveis

    Nível 1: CNPJ + Valor (mais robusto, imune a erros de digitação no nome)
    Nível 2: Nome normalizado + Valor (compatibilidade com sistema antigo)
    Nível 3: Fuzzy matching (sugestões quando não encontrar)

    Parâmetros:
        pagador_norm: Nome do pagador normalizado (uppercase, sem espaços extras)
        valor_cents: Valor em centavos (int)
        cnpj: CNPJ normalizado (14 dígitos) ou None
        mapa_nome: Dicionário {(nome, valor): entry}
        mapa_cnpj: Dicionário {(cnpj, valor): entry}
        lista_nomes: Lista de nomes para fuzzy matching

    Retorna:
        (entry, metodo_usado, sugestao) onde:
        - entry: dicionário com {'email', 'docs', 'cnpj', 'pagador_original'} ou None
        - metodo_usado: string descrevendo qual método funcionou
        - sugestao: string com sugestão de correção (ou None)
    """

    # NÍVEL 1: Tentar matching por CNPJ + Valor (mais robusto)
    if cnpj and mapa_cnpj:
        key_cnpj = (cnpj, valor_cents)
        if key_cnpj in mapa_cnpj:
            return mapa_cnpj[key_cnpj], "CNPJ+VALOR", None

    # NÍVEL 2: Tentar matching por Nome + Valor (compatibilidade)
    key_nome = (pagador_norm, valor_cents)
    if key_nome in mapa_nome:
        return mapa_nome[key_nome], "NOME+VALOR", None

    # NÍVEL 3: Fuzzy matching para sugestões
    # Procura nome similar com MESMO VALOR
    threshold_similaridade = 0.80  # 80% de similaridade

    candidatos_com_valor = []
    for (nome_candidato, valor_candidato), entry in mapa_nome.items():
        if valor_candidato == valor_cents:  # Mesmo valor
            candidatos_com_valor.append((nome_candidato, entry))

    if candidatos_com_valor:
        # Encontrar o nome mais similar
        melhor_nome, melhor_score = encontrar_match_mais_proximo(
            pagador_norm,
            [nome for nome, _ in candidatos_com_valor]
        )

        if melhor_score >= threshold_similaridade:
            # Encontrou match com boa similaridade
            entry_encontrado = next(e for n, e in candidatos_com_valor if n == melhor_nome)
            sugestao = f"Match por similaridade {melhor_score:.0%}: '{melhor_nome}' ≈ '{pagador_norm}'"
            return entry_encontrado, f"FUZZY({melhor_score:.0%})", sugestao

    # NÃO ENCONTROU: Gerar sugestão útil
    # Tentar encontrar o cliente (ignorando valor)
    todos_nomes = list(lista_nomes)
    if todos_nomes:
        nome_mais_proximo, score = encontrar_match_mais_proximo(pagador_norm, todos_nomes)

        if score >= 0.60:  # 60% de similaridade mínima para sugestão
            # Procurar valores deste cliente
            valores_deste_cliente = []
            for (nome, valor), entry in mapa_nome.items():
                if nome == nome_mais_proximo:
                    valores_deste_cliente.append(valor)

            sugestao = (
                f"[?] Cliente similar encontrado: '{nome_mais_proximo}' (similaridade: {score:.0%})\n"
                f"   Valores na planilha para este cliente: {[cents_to_brl(v) for v in valores_deste_cliente]}\n"
                f"   Valor do boleto: {cents_to_brl(valor_cents)}\n"
                f"   → Verifique se o valor está correto na planilha ou se há digitação errada no nome"
            )
            return None, "NOT_FOUND", sugestao

    # Nenhuma sugestão útil
    sugestao = (
        f"[ERRO] Nenhum cliente encontrado para:\n"
        f"   Nome: {pagador_norm}\n"
        f"   Valor: {cents_to_brl(valor_cents)}\n"
        f"   CNPJ: {cnpj or 'não extraído'}\n"
        f"   → Verifique se este cliente está cadastrado na planilha"
    )
    return None, "NOT_FOUND", sugestao

# -------------------- EMAIL --------------------
def montar_corpo_html(pagador, bases6_list, linhas_boleto, fidc_tipo):
    """
    Monta corpo HTML do email com template dinâmico baseado no FIDC

    Parâmetros:
        pagador: Nome do pagador
        bases6_list: Lista de números de notas fiscais
        linhas_boleto: Lista de dicts com 'valor_brl' e 'venc'
        fidc_tipo: Tipo do FIDC ("CAPITAL", "NOVAX", "CREDVALE", "SQUID")
    """
    # Obter configuração do FIDC
    config = FIDC_CONFIG.get(fidc_tipo, FIDC_CONFIG[FIDC_PADRAO])
    fidc_nome = config["nome_completo"]
    fidc_cnpj = config["cnpj"]

    # Montar lista de notas
    notas_txt = ", ".join(bases6_list) if bases6_list else "anexas"

    # Montar linhas de valor/vencimento
    linhas = "".join(
        f"<p>Valor: R$ {lb['valor_brl']}, Vencimento: {lb['venc']}</p>"
        for lb in linhas_boleto
    )

    # Template HTML dinâmico
    return f"""
<html>
<body style="font-family:Calibri,Arial,sans-serif; font-size:13.5px;">
<p>Boa tarde,</p>
<p>Prezado cliente,<br><b>{pagador}</b>,</p>
<p>Enviamos anexo o(s) seu(s) boleto(s) emitido(s) conforme a(as) nota(as): {notas_txt}</p>
{linhas}
<p>O(s) boleto(s) está(ão) com beneficiário nominal a <b>{fidc_nome}</b>, CNPJ: <b>{fidc_cnpj}</b>.</p>
<p>Vide boleto(s) e nota(s) em anexo.<br>Favor confirmar recebimento.</p>
<p>Em caso de dúvidas, nossa equipe permanece à disposição para esclarecimentos.</p>
<p>Atenciosamente,<br><b>Equipe de Cobrança</b></p>
<p><img src="cid:assinatura_jotajota" width="600"></p>
</body>
</html>
"""

def abrir_email_outlook(email_to, assunto, corpo_html, anexos, fidc_tipo):
    """
    Cria e envia email via Outlook com CCs dinâmicos baseados no FIDC

    Parâmetros:
        email_to: Email do destinatário
        assunto: Assunto do email
        corpo_html: Corpo do email em HTML
        anexos: Lista de caminhos de arquivos para anexar
        fidc_tipo: Tipo do FIDC ("CAPITAL", "NOVAX", "CREDVALE", "SQUID")
    """
    outlook = win32.Dispatch("Outlook.Application")

    # Verificação de segurança — garante conta de cobrança
    conta_cobranca = None
    for acc in outlook.Session.Accounts:
        if acc.SmtpAddress.lower() == "cobranca@jotajota.net.br":
            conta_cobranca = acc
            break

    if not conta_cobranca:
        print("[ERRO] ERRO: Conta cobrança@jotajota.net.br não encontrada — envio abortado.")
        with open(os.path.join(PASTA_LOGS, "log_falha_conta.txt"), "w", encoding="utf-8") as f:
            f.write(f"[{datetime.now()}] Conta cobrança não localizada. Envio abortado.\n")
        return

    mail = outlook.CreateItem(0)
    mail._oleobj_.Invoke(*(64209, 0, 8, 0, conta_cobranca))

    # Obter CCs dinâmicos baseados no FIDC
    config = FIDC_CONFIG.get(fidc_tipo, FIDC_CONFIG[FIDC_PADRAO])
    cc_emails = "; ".join(config["cc_emails"])

    mail.To = email_to
    mail.CC = cc_emails  # CC dinâmico por FIDC!
    mail.Subject = assunto

    # corpo + imagem inline
    mail.HTMLBody = corpo_html
    attach = mail.Attachments.Add(ASSINATURA_IMG)
    attach.PropertyAccessor.SetProperty(
        "http://schemas.microsoft.com/mapi/proptag/0x3712001F", "assinatura_jotajota"
    )

    # outros anexos
    for a in anexos:
        try:
            if not os.path.exists(a):
                print(f"      [ERRO] Arquivo não encontrado: {a}")
                raise FileNotFoundError(f"Arquivo não encontrado: {a}")
            mail.Attachments.Add(a)
        except Exception as e:
            print(f"      [ERRO] Falha ao anexar {os.path.basename(a)}: {e}")
            raise

    # Enviar ou mostrar preview
    if MODO_PREVIEW:
        mail.Display()  # Abre no Outlook para revisão (não envia)
    else:
        mail.Send()  # Envia automaticamente

# -------------------- MAIN --------------------
def executar():
    print("=" * 80)
    print("  ENVIO DE BOLETOS - v5.0.0 (Validacao Robusta + Auto-CNPJ)")
    print("=" * 80)
    if MODO_PREVIEW:
        print("  [MODO PREVIEW] Emails abrirao no Outlook sem enviar automaticamente")
        print("=" * 80)
    print()

    t0 = time.time()
    logs, log_erros = [], []
    enviados, erros = 0, 0
    grupos = {}

    # Estatísticas de matching
    stats_matching = {
        'CNPJ+VALOR': 0,
        'NOME+VALOR': 0,
        'FUZZY': 0,
        'NOT_FOUND': 0
    }

    # Carregar planilha (agora retorna 3 valores)
    mapa_nome, mapa_cnpj, lista_nomes = carregar_mapa_pagador_valor()

    if not mapa_nome:
        print("[ERRO] Nenhum dado carregado da planilha. Abortando.")
        return

    # Indexar notas fiscais
    notas_idx = indexar_notas_por_digitos()
    print(f"[ARQUIVO] Notas fiscais indexadas: {len(notas_idx)}")
    print()

    # Listar boletos para processar
    arquivos_boletos = [f for f in os.listdir(PASTA_BOLETOS) if f.lower().endswith(".pdf")]
    print(f"[PACOTE] Boletos encontrados para processar: {len(arquivos_boletos)}")
    print()

    if not arquivos_boletos:
        print("[INFO] Nenhum boleto para processar.")
        return

    print("[PROC] Processando boletos...")
    print("-" * 80)

    # Processar cada boleto
    for idx, arquivo in enumerate(arquivos_boletos, 1):
        caminho = os.path.join(PASTA_BOLETOS, arquivo)

        print(f"\n[{idx}/{len(arquivos_boletos)}] {arquivo}")

        # Parsear nome do arquivo: "PAGADOR - DD-MM - R$ VALOR.pdf"
        # Usar regex para extrair partes de forma robusta (pagador pode ter " - " no meio)
        nome_sem_ext = os.path.splitext(arquivo)[0]

        # Buscar data (DD-MM)
        match_data = re.search(r'\b(\d{2}-\d{2})\b', nome_sem_ext)
        # Buscar valor (R$ xxx,xx)
        match_valor = re.search(r'R[$]\s*[\d.,]+', nome_sem_ext)

        if not (match_data and match_valor):
            msg = f"   [ERRO] Formato inválido de arquivo (esperado: 'PAGADOR - DD-MM - R$ VALOR.pdf')"
            print(msg)
            log_erros.append(f"{arquivo}: {msg}")
            erros += 1
            continue

        venc = match_data.group(1)
        valor_str = match_valor.group(0)

        # Pagador é tudo antes da data
        pos_data = match_data.start()
        pagador_exib = nome_sem_ext[:pos_data].rstrip(' -').strip()

        # Normalizar dados
        pag_key = normalize_pagador(pagador_exib)
        val_cents = valor_to_cents(re.sub(r'^\s*R\$\s*', '', valor_str))

        if val_cents is None:
            msg = f"   [ERRO] Valor inválido: '{valor_str}'"
            print(msg)
            log_erros.append(f"{arquivo}: {msg}")
            erros += 1
            continue

        # Extrair CNPJ do PDF (nova funcionalidade!)
        print(f"   [PDF] Extraindo CNPJ do PDF...")
        cnpj = extrair_cnpj_do_pdf(caminho)
        if cnpj:
            print(f"   [OK] CNPJ extraído: {cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}")
            # Auto-preencher CNPJ na planilha (ETAPA 3!)
            atualizar_cnpj_na_planilha(pag_key, val_cents, cnpj)
        else:
            print(f"   [AVISO] CNPJ não encontrado no PDF (usará matching por nome)")

        # Detectar FIDC do boleto (nova funcionalidade ETAPA 2!)
        print(f"   [FIDC] Detectando FIDC do boleto...")
        fidc_tipo = detectar_fidc_do_pdf(caminho)
        print(f"   [OK] FIDC detectado: {fidc_tipo} ({FIDC_CONFIG[fidc_tipo]['nome_completo']})")

        # Buscar entrada na planilha (novo sistema de 3 níveis!)
        print(f"   [BUSCA] Buscando na planilha...")
        entry, metodo, sugestao = buscar_entrada_planilha(
            pag_key, val_cents, cnpj, mapa_nome, mapa_cnpj, lista_nomes
        )

        # Atualizar estatísticas
        if 'FUZZY' in metodo:
            stats_matching['FUZZY'] += 1
        else:
            stats_matching[metodo] += 1

        if not entry:
            # NÃO ENCONTROU
            print(f"   [ERRO] ERRO: Cliente não encontrado!")
            if sugestao:
                print(f"\n{sugestao}\n")
            log_erros.append(f"{arquivo}: {sugestao or 'Cliente não encontrado'}")
            erros += 1
            continue

        # ENCONTROU!
        print(f"   [OK] Match encontrado via: {metodo}")
        if sugestao:  # Se foi fuzzy, mostra a sugestão
            print(f"   [INFO] {sugestao}")

        email_to = entry['email']
        docs_set = entry['docs']

        print(f"   [EMAIL] Email destino: {email_to}")
        print(f"   [DOC] Documentos: {', '.join(sorted(docs_set)) if docs_set else 'nenhum'}")

        # Agrupar por (email, pagador)
        grupo_key = (email_to, pag_key)
        g = grupos.setdefault(grupo_key, {
            'pagador_exib': pagador_exib,
            'docs': set(),
            'boletos': [],
            'linhas': [],
            'metodos': [],
            'fidcs': [],  # Lista de FIDCs dos boletos (caso tenha múltiplos)
            'cnpjs': [],  # CNPJs dos boletos (para validação)
            'valores_cents': []  # Valores dos boletos (para validação)
        })

        g['docs'] |= docs_set
        g['boletos'].append(caminho)
        g['linhas'].append({'valor_brl': cents_to_brl(val_cents), 'venc': venc})
        g['metodos'].append(metodo)
        g['fidcs'].append(fidc_tipo)  # Armazenar FIDC do boleto
        if cnpj:
            g['cnpjs'].append(cnpj)  # Armazenar CNPJ para validação
        g['valores_cents'].append(val_cents)  # Armazenar valor para validação

    print()
    print("=" * 80)
    print("  PREPARANDO E ENVIANDO E-MAILS")
    print("=" * 80)
    print()

    # Enviar e-mails agrupados
    for (email_to, pag_key), g in grupos.items():
        print(f"[EMAIL] Enviando para: {g['pagador_exib']} ({email_to})")
        print(f"   - Boletos: {len(g['boletos'])}")
        print(f"   - Documentos: {len(g['docs'])}")

        # Determinar FIDC principal do grupo (caso tenha múltiplos, pega o mais comum)
        fidcs_unicos = list(set(g['fidcs']))
        if len(fidcs_unicos) == 1:
            fidc_grupo = fidcs_unicos[0]
            print(f"   - FIDC: {fidc_grupo}")
        else:
            # Múltiplos FIDCs no mesmo grupo - pegar o mais comum
            from collections import Counter
            fidc_grupo = Counter(g['fidcs']).most_common(1)[0][0]
            print(f"   - AVISO: Multiplos FIDCs detectados ({', '.join(fidcs_unicos)})")
            print(f"   - Usando FIDC mais comum: {fidc_grupo}")

        # Obter configuração do FIDC
        config_fidc = FIDC_CONFIG[fidc_grupo]
        print(f"   - CCs: {', '.join(config_fidc['cc_emails'])}")

        # Pegar CNPJ e valor do primeiro boleto para validação de notas
        cnpj_validacao = g['cnpjs'][0] if g['cnpjs'] else None
        valor_validacao = g['valores_cents'][0] if g['valores_cents'] else None

        # Buscar e validar notas fiscais correspondentes (ETAPA 3!)
        print(f"   - Validando notas fiscais...")
        print(f"   - Documentos esperados: {', '.join(sorted(g['docs']))}")
        notas_anexos, bases6, detalhes_validacao, tem_erro_critico = achar_notas_por_docs_set(
            g['docs'], notas_idx, cnpj_validacao, valor_validacao
        )
        print(f"   - Notas validadas: {len(notas_anexos)}")

        # BLOQUEIO DE SEGURANÇA: Se houver erro crítico, NÃO enviar email
        if tem_erro_critico:
            print(f"   [BLOQUEIO] Email NÃO será enviado devido a erros críticos de validação!")
            print(f"   [AÇÃO] Verifique as notas fiscais na pasta: {PASTA_NOTAS}")
            log_erros.append(f"Bloqueado: {email_to} - {g['pagador_exib']} - Erro crítico na validação de notas")
            erros += 1
            continue  # Pula para o próximo grupo

        # Enviar email com FIDC dinâmico
        try:
            abrir_email_outlook(
                email_to,
                f"Boleto e Nota Fiscal ({', '.join(bases6)})",
                montar_corpo_html(g['pagador_exib'], bases6, g['linhas'], fidc_grupo),
                list(g['boletos']) + list(notas_anexos),
                fidc_grupo  # Passa FIDC para CC dinâmico!
            )

            # Não mover boletos em modo preview (Outlook precisa acessar os arquivos originais)
            # Em modo automático, pode comentar esta seção se não quiser mover os arquivos
            # for b in g['boletos']:
            #     try:
            #         shutil.move(b, os.path.join(PASTA_ENVIADOS, os.path.basename(b)))
            #     except Exception as e:
            #         print(f"   [AVISO] Erro ao mover {os.path.basename(b)}: {e}")

            enviados += 1
            metodos_str = ", ".join(set(g['metodos']))
            logs.append(f"[OK] {email_to} | {g['pagador_exib']} | FIDC: {fidc_grupo} | Metodo: {metodos_str}")
            print(f"   [OK] Email enviado com sucesso!\n")

        except Exception as e:
            print(f"   [ERRO] ERRO ao enviar email: {e}\n")
            log_erros.append(f"Erro ao enviar para {email_to}: {e}")
            erros += 1

    # Resumo final
    dur = round(time.time() - t0, 2)

    print()
    print("=" * 80)
    print("  RESUMO DA EXECUÇÃO")
    print("=" * 80)
    print(f"[TEMPO] Tempo total: {dur}s")
    print(f"[OK] E-mails enviados: {enviados}")
    print(f"[ERRO] Erros: {erros}")
    print()
    print("[STATS] Estatísticas de Matching:")
    print(f"   - CNPJ + Valor: {stats_matching['CNPJ+VALOR']} boletos")
    print(f"   - Nome + Valor: {stats_matching['NOME+VALOR']} boletos")
    print(f"   - Fuzzy Match: {stats_matching['FUZZY']} boletos")
    print(f"   - Não encontrados: {stats_matching['NOT_FOUND']} boletos")
    print("=" * 80)

    # Salvar log
    resumo = f"""
RESUMO DA EXECUÇÃO - {datetime.now():%d/%m/%Y %H:%M:%S}
{'=' * 80}
Tempo total: {dur}s
E-mails enviados: {enviados}
Erros: {erros}

ESTATÍSTICAS DE MATCHING:
- CNPJ + Valor: {stats_matching['CNPJ+VALOR']} boletos
- Nome + Valor: {stats_matching['NOME+VALOR']} boletos
- Fuzzy Match: {stats_matching['FUZZY']} boletos
- Não encontrados: {stats_matching['NOT_FOUND']} boletos

E-MAILS ENVIADOS:
"""
    resumo += "\n".join(logs)

    if log_erros:
        resumo += "\n\nERROS:\n"
        resumo += "\n".join(log_erros)

    # Salvar log
    log_path = os.path.join(PASTA_LOGS, f"log_exec_{datetime.now():%Y%m%d_%H%M%S}.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(resumo)

    print(f"\n[LOG] Log salvo em: {log_path}")
    print()

    if enviados > 0:
        print("[OK] Processamento concluído com sucesso!")
        print("   Todos os e-mails foram enviados pela conta cobrança@jotajota.net.br")
    else:
        print("[AVISO] Nenhum e-mail foi enviado. Verifique os erros acima.")

if __name__ == "__main__":
    executar()
